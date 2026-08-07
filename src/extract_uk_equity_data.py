"""
Task 12 (internship Week 5): extract real UK share-level data from a Bloomberg Terminal export,
producing data/equities/uk_shares_returns_real.csv and data/equities/share_metadata_real.csv in the
exact shape equity_income.py already expects (see its load_equity_returns/load_share_metadata) - so
switching from placeholder to real data is a one-line change (point EQUITY_RETURNS_CSV/
SHARE_METADATA_CSV at the _real files, or just rename them over the placeholder ones) rather than a
rewrite of any downstream analysis.

------------------------------------------------------------------------------------------------
HOW TO PRODUCE THE BLOOMBERG EXPORT THIS SCRIPT EXPECTS (do this in Bloomberg Terminal / Excel first)
------------------------------------------------------------------------------------------------
This mirrors extract_data.py's existing convention exactly (same repeating-block layout as
Bloomberg_Decumulation_Data_9_July_2026.xlsx), so the same parsing logic can be reused unchanged:

1. Candidate universe: data/equities/uk_shares_universe_candidates.csv lists 16 real, long-listed
   FTSE 100 shares across 6 sectors (Consumer Staples, Healthcare, Financials, Utilities, Mining,
   Energy) - adjust the list first if you want different/more names, but keep enough sector spread
   for the basket search (task 14) to have genuine diversification to find.

2. For EACH ticker in that list, pull a MONTHLY TOTAL RETURN series (dividends reinvested, GBP,
   NOT just a bare price series - a price-only series would understate income shares' real return
   and defeat the point of testing "equities for retirement income"). In the Bloomberg Excel
   Add-in this is typically the monthly-periodicity BDH() pull on a total-return field (e.g.
   TOT_RETURN_INDEX_GROSS_DVDS), exactly as was used for the asset-class-level indices in
   Bloomberg_Decumulation_Data_9_July_2026.xlsx - ask whoever ran that original pull for the exact
   field/settings used there so this new pull is methodologically consistent with the rest of the
   model.

3. Pull as far back as Bloomberg has data for each name (ideally back to 1999-2000 to match the
   rest of this model's history; several of these companies will have less - that's fine, each
   share's own available history is used, exactly like the main app's per-portfolio history
   handling).

4. Lay the export out exactly like the existing Bloomberg file: one 3-column block per ticker
   (col 1 = Date, col 2 = Monthly Return, col 3 = blank spacer), row 1 = the ticker/company name as
   a block header, row 2 = column labels, data from row 3 down. Save as .xlsx.

5. Update SRC below to point at that file, then run: `python extract_uk_equity_data.py`

If you'd rather hand this off, this docstring + data/equities/uk_shares_universe_candidates.csv is
everything a colleague with terminal access needs to reproduce the pull.
------------------------------------------------------------------------------------------------
"""
import re
from pathlib import Path

import openpyxl
import pandas as pd

SRC = r"C:\Users\YahampathH\Downloads\bloomberg_monthly_returns_.xlsx"
SHEET_NAME = "Monthly Returns"  # this export's actual sheet name (template default), not "Bloomberg Direct Returns"

# DAY_TO_DAY_TOT_RETURN_GROSS_DVDS can come back in PERCENTAGE POINTS (e.g. -16.9045 meaning
# -16.9045%) or already in the decimal convention (-0.169045) the rest of this model uses (see
# data/equities/raise_index_returns.csv) DEPENDING ON HOW EACH CELL WAS PULLED - confirmed by
# inspecting one real export where one ticker's column was in percentage points and the other 15
# were already decimal, i.e. this is NOT consistent across a single file and can't be fixed with one
# fixed divisor. Detected and corrected per-column below instead (see _fix_scale).
PERCENT_MEDIAN_ABS_THRESHOLD = 1.0  # real monthly equity returns essentially never exceed 100% (1.0
                                     # in decimal terms), so a column whose typical |value| sits above
                                     # this is almost certainly in percentage points, not decimal

OUT_DIR = Path(__file__).resolve().parent.parent / "data" / "equities"
CANDIDATES_CSV = OUT_DIR / "uk_shares_universe_candidates.csv"


def extract(src=SRC, sheet_name=SHEET_NAME):
    """Same repeating-3-column-block parser as extract_data.py, applied to a share-level export
    instead of an asset-class-level one. Returns a wide DataFrame (Date index, one column per
    ticker/company name found in the export's block headers)."""
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[sheet_name]

    max_col = ws.max_column
    max_row = ws.max_row

    blocks = []
    col = 1
    while col <= max_col:
        header = ws.cell(row=1, column=col).value
        if header:
            # Guards against a stray typo'd header (e.g. "SVT LN Equityb") breaking the metadata
            # match below - truncates anything trailing immediately after "Equity" or "Index".
            clean = re.sub(r"((?:Equity|Index))\S*$", r"\1", str(header).strip())
            blocks.append((col, clean))
        col += 1

    print(f"Found {len(blocks)} data blocks in {src}")

    series = {}
    for start_col, name in blocks:
        dates, vals = [], []
        for r in range(3, max_row + 1):
            d = ws.cell(row=r, column=start_col).value
            v = ws.cell(row=r, column=start_col + 1).value
            if d is None:
                continue
            # Bloomberg errors (e.g. "#N/A") come back as strings - treated as missing, not coerced.
            vals.append(v if isinstance(v, (int, float)) else None)
            dates.append(pd.Timestamp(d))
        norm_dates = pd.DatetimeIndex(dates).to_period("M").to_timestamp("M")
        s = pd.Series(vals, index=norm_dates, name=name, dtype=float)
        s = s[~s.index.duplicated(keep="last")].sort_index()
        series[name] = s

    return _fix_scale(pd.DataFrame(series))


def _fix_scale(df: pd.DataFrame) -> pd.DataFrame:
    """Per-column: divides by 100 only if that column's own typical magnitude looks like percentage
    points rather than decimal (see PERCENT_MEDIAN_ABS_THRESHOLD) - handles a single export mixing
    both conventions across different columns, which a single fixed divisor cannot."""
    for col in df.columns:
        med_abs = df[col].abs().median()
        if med_abs > PERCENT_MEDIAN_ABS_THRESHOLD:
            df[col] = df[col] / 100.0
            print(f"  {col}: median |value| was {med_abs:.2f} -> treated as percentage points, divided by 100")
    return df


def match_metadata(columns) -> tuple[pd.DataFrame, dict]:
    """Matches extracted block-header names (e.g. "ULVR LN Equity") back to the candidate universe's
    short Ticker/Company/Sector so share_metadata_real.csv is populated automatically wherever the
    header naming lines up - anything left unmatched is printed so it can be fixed by hand. Also
    returns a {block_header: short_ticker} rename map so uk_shares_returns_real.csv's columns read
    as the short code ("ULVR") like every other file in the project (placeholder data, portfolio
    holdings, etc.), not the full Bloomberg security string."""
    candidates = pd.read_csv(CANDIDATES_CSV)
    rows = []
    rename = {}
    unmatched = []
    for col in columns:
        key = col.strip().upper()
        match = candidates[
            (candidates["Ticker"].str.upper() == key)
            | (candidates["BloombergTicker"].str.upper() == key)
            | (candidates["Company"].str.upper() == key)
        ]
        if len(match):
            row = match.iloc[0]
            rows.append({"Ticker": row["Ticker"], "Company": row["Company"], "Sector": row["Sector"]})
            rename[col] = row["Ticker"]
        else:
            unmatched.append(col)
            rows.append({"Ticker": col, "Company": col, "Sector": "Unknown - fix by hand"})
    if unmatched:
        print("Could not auto-match metadata for:", unmatched, "- edit share_metadata_real.csv by hand.")
    return pd.DataFrame(rows), rename


def main():
    if not Path(SRC).exists():
        print(
            f"SRC ({SRC}) not found - this script is ready to run but needs a real Bloomberg export "
            "first. See the docstring at the top of this file for exactly what to pull and how to "
            "lay it out."
        )
        return
    df = extract()
    meta, rename = match_metadata(df.columns)
    df = df.rename(columns=rename)

    returns_out = OUT_DIR / "uk_shares_returns_real.csv"
    meta_out = OUT_DIR / "share_metadata_real.csv"
    df.to_csv(returns_out)
    meta.to_csv(meta_out, index=False)
    print(f"Saved {returns_out} {df.shape} and {meta_out} {meta.shape}")
    print(
        "\nNext step: point equity_income.py's EQUITY_RETURNS_CSV/SHARE_METADATA_CSV at these two "
        "_real files (or rename them over the placeholder ones) and re-run "
        "run_equity_income_analysis.py - every downstream function (rank_shares, find_best_baskets, "
        "evaluate_basket, share_correlation_matrix) works unchanged."
    )


if __name__ == "__main__":
    main()
