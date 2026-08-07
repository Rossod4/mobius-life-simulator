"""
Builds output/Mobius_Wealth_UK_Equities_Visualiser.xlsx - a fully editable companion to
Mobius_Wealth_UK_Equities_Weeks5-6_Findings.docx (Task 17), covering the same ground (Task 13
individual-share ranking, Task 14 basket search, Task 16 rebalancing, share characteristics and
correlation) as NATIVE Excel charts and formulas rather than static images, so it can be reshaped,
recoloured, or extended directly in Excel rather than needing this script re-run.

House style matches the project's other Excel deliverables (build_workbook_*.py): Arial, dark-blue
header fill (#1F4E78), and (for the correlation sheet) the same red-white-blue ColorScaleRule
heatmap convention used in build_workbook_10.py's asset correlation sheet.

Run: `python build_equity_visualiser.py`
"""
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import ColorScaleRule

from engine import load_asset_returns, load_cpi, ClientProfile
from equity_income import (
    load_equity_returns, load_share_metadata, rank_shares, share_correlation_matrix,
    find_best_baskets, evaluate_basket,
)

OUT = Path(__file__).resolve().parent.parent / "output" / "Mobius_Wealth_UK_Equities_Visualiser.xlsx"

FONT = "Arial"
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1A2B3C")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="5A5A5A")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
BOLD = Font(name=FONT, bold=True)
BLACK = Font(name=FONT, color="000000")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = "0.00%"
GBP = '£#,##0;(£#,##0)'

SECTOR_HEX = {
    "Consumer Staples": "1BAF7A",
    "Healthcare": "3C7DC4",
    "Financials": "EDA100",
    "Utilities": "6B6F76",
    "Mining": "8B5E3C",
    "Energy": "D03B3B",
}


def header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ============================================================================================
# Compute results (same as build_equity_findings_report.py / run_equity_income_analysis.py)
# ============================================================================================
print("Computing results...")
asset_df = load_asset_returns()
cpi = load_cpi(asset_df)
equity_df = load_equity_returns()
meta = load_share_metadata()
tickers = list(meta["Ticker"])

profile = ClientProfile(starting_age=65, horizon_years=30, starting_pot=500_000.0, initial_annual_spend=20_000.0)

ranked = rank_shares(equity_df, cpi, profile)
ranked = ranked.merge(meta[["Ticker", "Company", "Sector"]], left_on="Share", right_on="Ticker")
ranked = ranked.sort_values("Probability of ruin").reset_index(drop=True)

sector_map = dict(zip(meta["Ticker"], meta["Sector"]))
top_baskets = find_best_baskets(equity_df, cpi, profile, basket_size=3, top_n=5,
                                 n_sims=300, sector_map=sector_map, min_sectors=3)

best_basket = top_baskets.iloc[0]["Basket"]
best_tickers = best_basket.split(" + ")
weights = {t: 1.0 / len(best_tickers) for t in best_tickers}
rebal_rows = []
for label, mode in [("Constant-mix (monthly)", "monthly"), ("Annual rebalance", "annual"),
                     ("Buy-and-hold", "buy_and_hold")]:
    res, dd = evaluate_basket(f"Best basket ({mode})", weights, equity_df, cpi, profile, rebalance=mode)
    s = res.summary()
    rebal_rows.append((label, s["Probability of ruin"], s["Median legacy"], dd["maxdd"]))

corr = share_correlation_matrix(equity_df[tickers])
print("Building workbook...")

# ============================================================================================
wb = openpyxl.Workbook()

# --- Overview ---
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 100
ws["B2"] = "Mobius Wealth — UK Equities for Retirement Income: Visualiser"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Companion to the Weeks 5-6 Findings write-up (Task 17) - editable charts, live formulas where practical."
ws["B3"].font = SUBTITLE_FONT
best_share = ranked.iloc[0]
ws["B5"] = (
    f"Best single share: {best_share['Share']} ({best_share['Company']}) - "
    f"{best_share['Probability of ruin']*100:.1f}% probability of ruin."
)
ws["B6"] = f"Best 3-share basket: {best_basket} - {top_baskets.iloc[0]['Probability of ruin']*100:.2f}% probability of ruin."
ws["B5"].font = BOLD
ws["B6"].font = BOLD
ws["B8"] = "Sheets:"
ws["B8"].font = BOLD
for i, (name, desc) in enumerate([
    ("Raw Returns", "Monthly total return, all 16 shares - edit or extend this and every formula elsewhere updates."),
    ("Characteristics", "Return/vol/sector table + scatter chart (colour = sector) + probability-of-ruin bar chart."),
    ("Correlation", "Full 16x16 CORREL() matrix with a live colour-scale heatmap (edit Raw Returns to see it update)."),
    ("Baskets & Rebalancing", "Task 14 basket comparison and Task 16 rebalancing comparison, as native bar charts."),
]):
    ws.cell(row=9 + i, column=2, value=f"  {name} — {desc}").font = BLACK

# --- Raw Returns ---
ws2 = wb.create_sheet("Raw Returns")
ws2.sheet_view.showGridLines = False
header_row(ws2, 1, ["Date"] + tickers)
for r, (date, row) in enumerate(equity_df[tickers].iterrows(), start=2):
    ws2.cell(row=r, column=1, value=date.strftime("%Y-%m-%d"))
    for c, t in enumerate(tickers, start=2):
        v = row[t]
        cell = ws2.cell(row=r, column=c, value=None if pd.isna(v) else float(v))
        cell.number_format = "0.00%"
n_data_rows = len(equity_df)
autosize(ws2, {"A": 12, **{get_column_letter(c): 9 for c in range(2, 2 + len(tickers))}})
ws2.freeze_panes = "B2"

# --- Characteristics ---
ws3 = wb.create_sheet("Characteristics")
ws3.sheet_view.showGridLines = False
ws3["B1"] = "Share characteristics"
ws3["B1"].font = TITLE_FONT
header_row(ws3, 3, ["Ticker", "Company", "Sector", "Ann. return", "Ann. vol", "Prob. of ruin", "Median legacy"], col_start=2)
first_data_row = 4
rank_by_ticker = ranked.set_index("Share")
last_data_row = first_data_row + len(tickers) - 1
autosize(ws3, {"B": 10, "C": 26, "D": 18, "E": 12, "F": 12, "G": 14, "H": 16})

# Rows are laid out sector-by-sector (not ticker order) so each sector's rows are a CONTIGUOUS
# range - needed to build one coloured scatter series per sector below (a single Reference range
# can't cherry-pick non-contiguous rows).
sector_order = sorted(tickers, key=lambda t: (list(SECTOR_HEX).index(sector_map[t]), t))
for i, t in enumerate(sector_order):
    r = first_data_row + i
    m = meta[meta["Ticker"] == t].iloc[0]
    ws3.cell(row=r, column=2, value=t).font = BLACK
    ws3.cell(row=r, column=3, value=m["Company"]).font = BLACK
    ws3.cell(row=r, column=4, value=m["Sector"]).font = BLACK
    col_letter = get_column_letter(2 + tickers.index(t))
    ret_cell = ws3.cell(row=r, column=5,
        value=f"=(1+AVERAGE('Raw Returns'!{col_letter}2:{col_letter}{1+n_data_rows}))^12-1")
    ret_cell.number_format = PCT
    vol_cell = ws3.cell(row=r, column=6,
        value=f"=STDEV('Raw Returns'!{col_letter}2:{col_letter}{1+n_data_rows})*SQRT(12)")
    vol_cell.number_format = PCT
    rr = rank_by_ticker.loc[t]
    pr_cell = ws3.cell(row=r, column=7, value=float(rr["Probability of ruin"]))
    pr_cell.number_format = PCT
    ml_cell = ws3.cell(row=r, column=8, value=float(rr["Median legacy"]))
    ml_cell.number_format = GBP
    for col in range(2, 9):
        ws3.cell(row=r, column=col).border = BORDER

scatter = ScatterChart()
scatter.title = "Return vs volatility, by sector"
scatter.x_axis.title = "Annualised volatility"
scatter.y_axis.title = "Annualised return"
scatter.x_axis.numFmt = "0%"
scatter.y_axis.numFmt = "0%"
scatter.height, scatter.width = 11, 18
for sector, hexcol in SECTOR_HEX.items():
    idxs = [i for i, t in enumerate(sector_order) if sector_map[t] == sector]
    if not idxs:
        continue
    r0, r1 = first_data_row + min(idxs), first_data_row + max(idxs)
    xref = Reference(ws3, min_col=6, min_row=r0, max_row=r1)
    yref = Reference(ws3, min_col=5, min_row=r0, max_row=r1)
    series = Series(yref, xref, title=sector)
    series.marker = Marker(symbol="circle", size=8)
    series.marker.graphicalProperties = GraphicalProperties(solidFill=hexcol)
    series.graphicalProperties.line.noFill = True
    scatter.series.append(series)
ws3.add_chart(scatter, "J3")

# Bar chart: probability of ruin per share, sorted, coloured per-sector via per-point fill
sorted_rows = list(range(first_data_row, last_data_row + 1))
sorted_rows.sort(key=lambda r: ws3.cell(row=r, column=7).value)
bar = BarChart()
bar.type = "bar"
bar.title = "Probability of ruin by share"
bar.y_axis.title = None
bar.x_axis.title = "Probability of ruin"
bar.x_axis.numFmt = "0%"
bar.height, bar.width = 11, 18
data_ref = Reference(ws3, min_col=7, min_row=first_data_row - 1, max_row=last_data_row)
cats_ref = Reference(ws3, min_col=2, min_row=first_data_row, max_row=last_data_row)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.legend = None
s = bar.series[0]
s.graphicalProperties.solidFill = "1F4E78"
s.dPt = [
    DataPoint(idx=i, spPr=GraphicalProperties(solidFill=SECTOR_HEX[sector_map[ws3.cell(row=first_data_row + i, column=2).value]]))
    for i in range(len(tickers))
]
ws3.add_chart(bar, "J24")

# --- Correlation ---
ws4 = wb.create_sheet("Correlation")
ws4.sheet_view.showGridLines = False
ws4["B1"] = "Pairwise correlation (live CORREL formulas - edit Raw Returns and this updates)"
ws4["B1"].font = TITLE_FONT
first_corr_row, first_corr_col = 3, 2
header_row(ws4, first_corr_row, [""] + tickers, col_start=first_corr_col)
for i, t_row in enumerate(tickers):
    r = first_corr_row + 1 + i
    label_cell = ws4.cell(row=r, column=first_corr_col, value=t_row)
    label_cell.font = HEADER_FONT
    label_cell.fill = HEADER_FILL
    row_col_letter = get_column_letter(2 + tickers.index(t_row))
    for j, t_col in enumerate(tickers):
        c = first_corr_col + 1 + j
        col_col_letter = get_column_letter(2 + tickers.index(t_col))
        if t_row == t_col:
            formula = 1.0
        else:
            formula = (f"=CORREL('Raw Returns'!${row_col_letter}$2:${row_col_letter}${1+n_data_rows},"
                       f"'Raw Returns'!${col_col_letter}$2:${col_col_letter}${1+n_data_rows})")
        cell = ws4.cell(row=r, column=c, value=formula)
        cell.number_format = "0.00"
        cell.font = BLACK
last_corr_row = first_corr_row + len(tickers)
last_corr_col = first_corr_col + len(tickers)
rule = ColorScaleRule(
    start_type="num", start_value=-1, start_color="D6604D",
    mid_type="num", mid_value=0, mid_color="FFFFFF",
    end_type="num", end_value=1, end_color="4393C3",
)
data_range = (f"{get_column_letter(first_corr_col+1)}{first_corr_row+1}:"
              f"{get_column_letter(last_corr_col)}{last_corr_row}")
ws4.conditional_formatting.add(data_range, rule)
autosize(ws4, {"A": 2, "B": 10, **{get_column_letter(c): 7 for c in range(3, last_corr_col + 1)}})

# --- Baskets & Rebalancing ---
ws5 = wb.create_sheet("Baskets & Rebalancing")
ws5.sheet_view.showGridLines = False
ws5["B1"] = "Task 14 — top 5 baskets found"
ws5["B1"].font = TITLE_FONT
header_row(ws5, 3, ["Basket", "Prob. of ruin", "Median legacy", "Max DD", "Avg DD", "Sectors"], col_start=2)
for i, (_, row) in enumerate(top_baskets.iterrows()):
    r = 4 + i
    ws5.cell(row=r, column=2, value=row["Basket"]).font = BLACK
    c = ws5.cell(row=r, column=3, value=float(row["Probability of ruin"])); c.number_format = PCT
    c = ws5.cell(row=r, column=4, value=float(row["Median legacy"])); c.number_format = GBP
    c = ws5.cell(row=r, column=5, value=float(row["Max DD"])); c.number_format = PCT
    c = ws5.cell(row=r, column=6, value=float(row["Average DD"])); c.number_format = PCT
    ws5.cell(row=r, column=7, value=row["Sectors"]).font = BLACK
    for col in range(2, 8):
        ws5.cell(row=r, column=col).border = BORDER
basket_last_row = 3 + len(top_baskets)

bar2 = BarChart()
bar2.type = "bar"
bar2.title = "Top 5 baskets - probability of ruin"
bar2.x_axis.numFmt = "0.0%"
bar2.height, bar2.width = 9, 16
data_ref2 = Reference(ws5, min_col=3, min_row=3, max_row=basket_last_row)
cats_ref2 = Reference(ws5, min_col=2, min_row=4, max_row=basket_last_row)
bar2.add_data(data_ref2, titles_from_data=True)
bar2.set_categories(cats_ref2)
bar2.legend = None
bar2.series[0].graphicalProperties.solidFill = "1BAF7A"
ws5.add_chart(bar2, "I3")

rebal_start = basket_last_row + 3
ws5.cell(row=rebal_start, column=2, value="Task 16 — rebalancing comparison for the winning basket").font = TITLE_FONT
header_row(ws5, rebal_start + 2, ["Approach", "Prob. of ruin", "Median legacy", "Max DD"], col_start=2)
for i, (label, p, ml, dd) in enumerate(rebal_rows):
    r = rebal_start + 3 + i
    ws5.cell(row=r, column=2, value=label).font = BLACK
    c = ws5.cell(row=r, column=3, value=float(p)); c.number_format = PCT
    c = ws5.cell(row=r, column=4, value=float(ml)); c.number_format = GBP
    c = ws5.cell(row=r, column=5, value=float(dd)); c.number_format = PCT
    for col in range(2, 6):
        ws5.cell(row=r, column=col).border = BORDER
rebal_last_row = rebal_start + 2 + len(rebal_rows)

bar3 = BarChart()
bar3.type = "col"
bar3.title = "Rebalancing convention vs probability of ruin"
bar3.y_axis.numFmt = "0.00%"
bar3.height, bar3.width = 9, 16
data_ref3 = Reference(ws5, min_col=3, min_row=rebal_start + 2, max_row=rebal_last_row)
cats_ref3 = Reference(ws5, min_col=2, min_row=rebal_start + 3, max_row=rebal_last_row)
bar3.add_data(data_ref3, titles_from_data=True)
bar3.set_categories(cats_ref3)
bar3.legend = None
bar3.series[0].graphicalProperties.solidFill = "3C7DC4"
ws5.add_chart(bar3, f"I{rebal_start}")

autosize(ws5, {"B": 30, "C": 13, "D": 15, "E": 11, "F": 11, "G": 40})

wb.save(OUT)
print(f"Saved {OUT}")
